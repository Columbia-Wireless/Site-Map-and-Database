"""
import_real_data.py
-------------------
Reads the four SCETV Excel attachments and generates INSERT SQL
for state_agencies, licensees, tower_sites, and site_licenses.

Usage:
  python scripts/import_real_data.py > scripts/import_data.sql

Then paste into the Supabase SQL editor, or run via psql.
"""

import openpyxl
import re
import sys
from datetime import datetime, date

# ── File paths ────────────────────────────────────────────────────────────────
ATT1 = "C:/Users/tomso/Downloads/Attachment 1_SCETV Towers.xlsx"
ATT2 = "C:/Users/tomso/Downloads/Attachment 2_License Count.xlsx"
ATT3 = "C:/Users/tomso/Downloads/Attachment 3_Buildings.xlsx"
ATT4 = "C:/Users/tomso/Downloads/Attachment 4_Land.xlsx"

def esc(v):
    """Escape a value for SQL: returns NULL or a quoted string."""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if s == "" or s.upper() in ("UNK", "UNKNOWN", "N/A", "NA", "NONE"):
        return "NULL"
    return "'" + s.replace("'", "''") + "'"

def esc_num(v):
    try:
        return str(float(str(v).strip().replace(",", "")))
    except:
        return "NULL"

def esc_date(v):
    if v is None:
        return "NULL"
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    if not s or s.upper() in ("UNK","N/A","NA","NONE","UNLIMITED 3 YR RENEWALS"):
        return "NULL"
    # Try to parse common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return f"'{datetime.strptime(s, fmt).strftime('%Y-%m-%d')}'"
        except:
            pass
    return "NULL"

# ── Agency type mapping ───────────────────────────────────────────────────────
AGENCY_NORMALIZE = {
    "CCU": "COASTAL CAROLINA UNIVERSITY",
    "DEPT OF PUBLIC SAFETY": "DEPARTMENT OF PUBLIC SAFETY",
    "Department of Transportation": "DEPARTMENT OF TRANSPORTATION",
    "USC FOUNDATIONS": "UNIVERSITY OF SOUTH CAROLINA - COLUMBIA CAMPUS",
    "University of South Carolina Lancaster": "UNIVERSITY OF SOUTH CAROLINA LANCASTER",
    "SC GOVERNOR'S SCHOOL FOR SCIENCE & MATH": "SC Governor's School for Science and Mathematics",
    "LAW ENFORCEMENT TRAINING COUNSEL - CRIMINAL JUSTICE ACADEMY": "LAW ENFORCEMENT TRAINING COUNCIL (CRIMINAL JUSTICE ACADEMY)",
    "DEPARTMENT OF LABOR  LICENSING AND REGULATIONS": "DEPARTMENT OF LABOR, LICENSING AND REGULATIONS",
    "DEPARTMENT OF PARKS  RECREATION AND TOURISM": "DEPARTMENT OF PARKS, RECREATION AND TOURISM",
    "USA": "other",
    "SPARTANBURG COMMUNITY COLLEGE FOUNDATION": "SPARTANBURG COMMUNITY COLLEGE",
    "STATE EDUCATIONAL FINANCE COMM": "DEPARTMENT OF EDUCATION",
    "TECHNICAL COLLEGE": "other",  # too generic - skip
}

AGENCY_TYPE_MAP = {
    "EDUCATIONAL TELEVISION COMMISSION": "broadcast",
    "EDUCATIONAL TELEVISION": "broadcast",
    "STATE LAW ENFORCEMENT DIVISION": "law_enforcement",
    "SLED": "law_enforcement",
    "DEPARTMENT OF ADMINISTRATION": "state_agency",
    "DEPARTMENT OF COMMERCE": "state_agency",
    "DEPARTMENT OF CORRECTIONS": "state_agency",
    "DEPARTMENT OF EDUCATION": "state_agency",
    "DEPARTMENT OF NATURAL RESOURCES": "state_agency",
    "DEPARTMENT OF JUVENILE JUSTICE": "state_agency",
    "DEPARTMENT OF TRANSPORTATION": "state_agency",
    "DEPARTMENT OF PUBLIC SAFETY": "state_agency",
    "DEPARTMENT OF PUBLIC HEALTH": "state_agency",
    "DEPARTMENT OF AGRICULTURE": "state_agency",
    "DEPARTMENT OF VETERANS AFFAIRS": "state_agency",
    "DEPARTMENT OF MOTOR VEHICLES": "state_agency",
    "DEPARTMENT OF BEHAVIORAL HEALTH AND DEVELOPMENTAL DISABILITIES": "state_agency",
    "DEPARTMENT OF EMPLOYMENT AND WORKFORCE": "state_agency",
    "DEPARTMENT OF ENVIRONMENTAL SERVICES": "state_agency",
    "DEPARTMENT OF PARKS, RECREATION AND TOURISM": "state_agency",
    "DEPARTMENT OF LABOR, LICENSING AND REGULATIONS": "state_agency",
    "FORESTRY COMMISSION": "state_agency",
    "AERONAUTICS COMMISSION": "state_agency",
    "COMMISSION FOR THE BLIND": "state_agency",
    "EMERGENCY MANAGEMENT DIVISION": "state_agency",
    "OFFICE OF THE ADJUTANT GENERAL": "military",
    "THE CITADEL": "university",
    "CITADEL": "university",
    "CLEMSON UNIVERSITY": "university",
    "COLLEGE OF CHARLESTON": "university",
    "COASTAL CAROLINA UNIVERSITY": "university",
    "FRANCIS MARION UNIVERSITY": "university",
    "LANDER UNIVERSITY": "university",
    "SOUTH CAROLINA STATE UNIVERSITY": "university",
    "WINTHROP UNIVERSITY": "university",
    "UNIVERSITY OF SOUTH CAROLINA - COLUMBIA CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - AIKEN CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - BEAUFORT CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - SALKEHATCHIE CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - SUMTER CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - UNION CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA - UPSTATE CAMPUS": "university",
    "UNIVERSITY OF SOUTH CAROLINA LANCASTER": "university",
    "MEDICAL UNIVERSITY OF SOUTH CAROLINA": "medical",
    "PUBLIC EMPLOYEE BENEFITS AUTHORITY": "state_agency",
    "AIKEN TECHNICAL COLLEGE": "technical_college",
    "CENTRAL CAROLINA TECHNICAL COLLEGE": "technical_college",
    "DENMARK TECHNICAL COLLEGE": "technical_college",
    "FLORENCE-DARLINGTON TECHNICAL COLLEGE": "technical_college",
    "GREENVILLE TECHNICAL COLLEGE": "technical_college",
    "HORRY-GEORGETOWN TECHNICAL COLLEGE": "technical_college",
    "MIDLANDS TECHNICAL COLLEGE": "technical_college",
    "NORTHEASTERN TECHNICAL COLLEGE": "technical_college",
    "ORANGEBURG-CALHOUN TECHNICAL COLLEGE": "technical_college",
    "PIEDMONT TECHNICAL COLLEGE": "technical_college",
    "SPARTANBURG COMMUNITY COLLEGE": "technical_college",
    "TECHNICAL COLLEGE OF THE LOWCOUNTRY": "technical_college",
    "TRIDENT TECHNICAL COLLEGE": "technical_college",
    "YORK TECHNICAL COLLEGE": "technical_college",
}

LICENSEE_TYPE_MAP = {
    "T-MOBILE": "carrier",
    "T-Mobile": "carrier",
    "VERIZON": "carrier",
    "AT&T": "carrier",
    "SPRINT": "carrier",
    "HORRY TELEPHONE": "carrier",
    "FARMER'S TELEPHONE": "carrier",
    "FBI": "government",
    "SLED": "government",
    "SC DNR": "government",
    "SANTEE COOPER": "government",
    "SC DOA": "government",
    "NOAA": "government",
    "FAA": "government",
    "U.S. FEDERAL": "government",
    "U.S. CUSTOMS": "government",
    "CHEROKEE COUNTY": "government",
    "K-LOVE": "broadcast",
    "SINCLAIR": "broadcast",
    "WCCP": "broadcast",
    "GRAY MEDIA": "broadcast",
    "WMBF": "broadcast",
    "WLOS": "broadcast",
}

TOWER_TYPE_MAP = {
    "Lattice": "lattice",
    "Monopole": "monopole",
    "Rooftop": "rooftop",
    "Water Tower": "water_tower",
    None: "other",
    "None": "other",
}

TOWER_STATUS_MAP = {
    "Active": "operational",
    "No Tower": "no_tower",
    "Removed": "removed",
    None: "operational",
    "None": "operational",
}

def get_licensee_type(name):
    if not name:
        return "other"
    nu = name.upper()
    for k, v in LICENSEE_TYPE_MAP.items():
        if k.upper() in nu:
            return v
    return "other"

def get_agency_type(name):
    if not name:
        return "other"
    nu = name.strip().upper()
    if nu in AGENCY_TYPE_MAP:
        return AGENCY_TYPE_MAP[nu]
    for k, v in AGENCY_TYPE_MAP.items():
        if k.upper() in nu or nu in k.upper():
            return v
    return "other"

# ── Parse license period string → (start, end) ───────────────────────────────
MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12
}

def parse_license_period(s):
    """Parse 'Jan 1, 2025- Dec 31, 2025' → ('2025-01-01', '2025-12-31')"""
    if not s:
        return None, None
    s = str(s).strip()
    parts = re.split(r"\s*-\s*", s, maxsplit=1)
    if len(parts) != 2:
        return None, None
    def parse_part(p):
        p = p.strip().rstrip(",").rstrip(".")
        # "Jan 1, 2025" or "July 31, 2026"
        m = re.match(r"([A-Za-z]+)\s+(\d+),?\s+(\d{4})", p)
        if m:
            mon = MONTH_MAP.get(m.group(1).lower()[:4])
            if not mon:
                mon = MONTH_MAP.get(m.group(1).lower()[:3])
            if mon:
                return f"{m.group(3)}-{mon:02d}-{int(m.group(2)):02d}"
        return None
    return parse_part(parts[0]), parse_part(parts[1])

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Collect unique state agencies from Attachment 3 + 4
# ─────────────────────────────────────────────────────────────────────────────
agencies = {}  # normalized_name -> {name, type, website}

# From buildings
wb3 = openpyxl.load_workbook(ATT3, read_only=True, data_only=True)
for sheet in wb3.sheetnames:
    ws = wb3[sheet]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if row[6]:
            name = str(row[6]).strip()
            # Normalize duplicates/aliases
            name = AGENCY_NORMALIZE.get(name, AGENCY_NORMALIZE.get(name.upper(), name))
            if not name or name == 'other':
                continue
            key = name.upper()
            if key not in agencies:
                website = str(row[24]).strip() if len(row) > 24 and row[24] else None
                agencies[key] = {
                    "name": name,
                    "type": get_agency_type(name),
                    "website": website if website and website.upper() not in ("NONE","N/A") else None,
                }
wb3.close()

# Also add ETC from Attachment 4 (land owner)
wb4 = openpyxl.load_workbook(ATT4, read_only=True, data_only=True)
for sheet in wb4.sheetnames:
    ws = wb4[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    for row in rows[1:]:
        if row[6]:
            name = str(row[6]).strip()
            name = AGENCY_NORMALIZE.get(name, AGENCY_NORMALIZE.get(name.upper(), name))
            if not name or name == 'other':
                continue
            key = name.upper()
            if key not in agencies:
                agencies[key] = {
                    "name": name,
                    "type": get_agency_type(name),
                    "website": None,
                }
wb4.close()

# Build land ownership map: address_upper → {agency_name, land_use}
land_tower_map = {}  # address_upper → (agency_key, land_use, parcel_ref)
wb4b = openpyxl.load_workbook(ATT4, read_only=True, data_only=True)
for sheet in wb4b.sheetnames:
    ws = wb4b[sheet]
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    for row in rows[1:]:
        if row[1]:  # address
            addr = str(row[1]).strip().upper()
            agency = str(row[6]).strip() if row[6] else ""
            land_use = str(row[14]).strip() if row[14] else ""
            parcel_ref = str(row[8]).strip() if row[8] else ""
            land_tower_map[addr] = (agency.upper(), land_use, parcel_ref)
wb4b.close()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Parse Attachment 1 tower sites
# ─────────────────────────────────────────────────────────────────────────────
wb1 = openpyxl.load_workbook(ATT1, read_only=True, data_only=True)
ws1 = wb1['Sheet1']
tower_rows = [r for r in ws1.iter_rows(values_only=True) if any(c is not None for c in r)][1:]
wb1.close()

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Parse Attachment 2 licensees + licenses
# ─────────────────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(ATT2, read_only=True, data_only=True)
ws2 = wb2['Sheet1']
lic_rows = [r for r in ws2.iter_rows(values_only=True) if any(c is not None for c in r)][1:]
wb2.close()

# Unique licensees
licensees = {}  # normalized_name → {name, type}
for row in lic_rows:
    if row[2]:
        name = str(row[2]).strip()
        key = name.upper()
        if key not in licensees:
            licensees[key] = {
                "name": name,
                "type": get_licensee_type(name),
            }

# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT SQL
# ─────────────────────────────────────────────────────────────────────────────
out = []
out.append("-- ============================================================")
out.append("-- SCETV Tower Management - Real Data Import")
out.append(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
out.append("-- ============================================================")
out.append("")
out.append("BEGIN;")
out.append("")

# ── State agencies ────────────────────────────────────────────────────────────
out.append("-- STATE AGENCIES")
out.append("INSERT INTO state_agencies (name, type, status) VALUES")
agency_vals = []
agency_name_to_idx = {}  # normalized → position in INSERT (1-based, for reference)

sorted_agencies = sorted(agencies.items(), key=lambda x: x[1]["name"])
for i, (key, a) in enumerate(sorted_agencies):
    if not a['name'] or a['name'].strip() == '':
        continue
    agency_name_to_idx[key] = i + 1
    agency_vals.append(
        f"  ({esc(a['name'])}, {esc(a['type'])}, 'active')"
    )
out.append(",\n".join(agency_vals) + ";")
out.append("")

# ── Licensees ─────────────────────────────────────────────────────────────────
out.append("-- LICENSEES")
out.append("INSERT INTO licensees (name, licensee_type, status) VALUES")
lic_vals = []
sorted_licensees = sorted(licensees.items(), key=lambda x: x[1]["name"])
for key, l in sorted_licensees:
    lic_vals.append(
        f"  ({esc(l['name'])}, {esc(l['type'])}, 'active')"
    )
out.append(",\n".join(lic_vals) + ";")
out.append("")

# ── Tower sites ───────────────────────────────────────────────────────────────
out.append("-- TOWER SITES (501 sites from Attachment 1)")
out.append("""INSERT INTO tower_sites
  (site_code, name, address, city, state, zip, county,
   lat, lng, tower_type, height_ft, status,
   host_agency_id, land_ownership, land_parcel_ref)
VALUES""")

site_vals = []
for i, r in enumerate(tower_rows):
    # Skip fully empty rows
    if not r[1] and not r[2]:
        continue
    # r: (row_num, title, address, city, zip, county, status, tower_type, height, lat, lng, ...)
    row_num = str(r[0]).strip() if r[0] else str(i+1)
    site_code = f"SC{int(float(row_num)):04d}" if row_num.replace(".","").isdigit() else f"SC{i+1:04d}"
    name     = str(r[1]).strip() if r[1] else f"Site {site_code}"
    address  = str(r[2]).strip() if r[2] else ""
    city     = str(r[3]).strip() if r[3] else ""
    zipcode  = str(r[4]).strip() if r[4] else ""
    county   = str(r[5]).strip() if r[5] else ""
    t_status = TOWER_STATUS_MAP.get(str(r[6]).strip() if r[6] else None, "operational")
    t_type   = TOWER_TYPE_MAP.get(str(r[7]).strip() if r[7] else None, "other")
    height   = int(float(str(r[8]))) if r[8] and str(r[8]).replace(".","").isdigit() else None
    lat      = float(r[9]) if r[9] else None
    lng      = float(r[10]) if r[10] else None

    # Determine land_ownership from Attachment 4 matches
    addr_upper = address.upper()
    land_own   = "unknown"
    parcel_ref = None
    host_agency_subq = "NULL"

    if addr_upper in land_tower_map:
        agency_key, land_use, parcel = land_tower_map[addr_upper]
        parcel_ref = parcel if parcel else None
        if "EDUCATIONAL TELEVISION" in agency_key or "ETC" == agency_key:
            land_own = "etc_owned"
        else:
            land_own = "other_agency"
        # Lookup agency id
        if agency_key in agencies:
            host_agency_subq = f"(SELECT id FROM state_agencies WHERE name = {esc(agencies[agency_key]['name'])} LIMIT 1)"

    height_sql = str(height) if height is not None else "NULL"
    lat_sql    = str(lat) if lat else "NULL"
    lng_sql    = str(lng) if lng else "NULL"

    site_vals.append(
        f"  ({esc(site_code)}, {esc(name)}, {esc(address)}, {esc(city)}, 'SC', {esc(zipcode)}, "
        f"{esc(county)}, {lat_sql}, {lng_sql}, {esc(t_type)}, {height_sql}, {esc(t_status)}, "
        f"{host_agency_subq}, {esc(land_own)}, {esc(parcel_ref)})"
    )

out.append(",\n".join(site_vals) + ";")
out.append("")

# ── Site licenses ─────────────────────────────────────────────────────────────
out.append("-- SITE LICENSES (69 licenses from Attachment 2)")
out.append("""INSERT INTO site_licenses
  (site_id, licensee_id, license_start, license_end, final_expiration, status, mount_type)
VALUES""")

lic_insert_vals = []
for row in lic_rows:
    # row: (count, title, licensee_name, period_str, expiration)
    title    = str(row[1]).strip() if row[1] else ""
    lic_name = str(row[2]).strip() if row[2] else ""
    period   = str(row[3]).strip() if row[3] else ""
    expiry   = row[4]

    ls, le = parse_license_period(period)
    ls_sql = f"'{ls}'" if ls else "NULL"
    le_sql = f"'{le}'" if le else "NULL"
    exp_sql = esc_date(expiry)

    # Determine status from license_end
    status = "active"
    if le:
        try:
            end_dt = datetime.strptime(le, "%Y-%m-%d").date()
            if end_dt < date.today():
                status = "expired"
        except:
            pass

    site_subq = f"(SELECT id FROM tower_sites WHERE UPPER(name) = UPPER({esc(title)}) LIMIT 1)"
    lic_subq  = f"(SELECT id FROM licensees WHERE UPPER(name) = UPPER({esc(lic_name)}) LIMIT 1)"

    lic_insert_vals.append(
        f"  ({site_subq}, {lic_subq}, {ls_sql}, {le_sql}, {exp_sql}, {esc(status)}, 'Primary')"
    )

out.append(",\n".join(lic_insert_vals) + ";")
out.append("")
out.append("COMMIT;")
out.append("")
out.append(f"-- Summary: {len(sorted_agencies)} agencies, {len(sorted_licensees)} licensees,")
out.append(f"--          {len(tower_rows)} tower sites, {len(lic_rows)} licenses")

print("\n".join(out))
